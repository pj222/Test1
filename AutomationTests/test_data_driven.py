import requests
import json
import jsonpath
import openpyxl


def test_addstud():
    #global id
    url="https://thetestingworldapi.com/api/studentsDetails"
    f=open("C:/Python/add.json","r")
    wk=openpyxl.load_workbook('C:/Python/add_data.xlsx')
    sh=wk['Sheet1']
    rows=sh.max_row
    json_request=json.loads(f.read())
    for i in range(2,rows+1):
        cell_fname=sh.cell(row=i,column=1)
        cell_mname = sh.cell(row=i, column=2)
        cell_lname = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)
        json_request['first_name']=cell_fname.value
        json_request['last_name'] = cell_lname.value
        json_request['middle_name'] = cell_mname.value
        json_request['date_of_birth'] = cell_dob.value
        json_response=requests.post(url,json_request)
        print(json_response.text)
        print(json_response.status_code)




   # print(id[0])
#def test_getdata():
 #   url = "https://thetestingworldapi.com/api/studentsDetails/"+str(id[0])
  #  req=requests.get(url)
   # print(req.text)
