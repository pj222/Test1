import requests
import json
import jsonpath
import openpyxl

class common:

    def __init__(self,filenamepath,sheetname):
        global wk
        global sh
        wk=openpyxl.load_workbook(filenamepath)
        sh=wk[sheetname]

    def fetch_row(self):
        rows=sh.max_row
        return rows

    def fetch_cols(self):
        cols=sh.max_column
        return cols

    def fetch_keynames(self):
        c=sh.max_column
        li=[]
        for i in range(1,c+1):
            cell=sh.cell(row=1,column=i)
            li.insert(i-1,cell.value)
            return li

    def insertvalue(self,rownumber,jsonrequest,keylist):
        c=sh.max_column
        for i in range(1,c+1):
            cell=sh.cell(row=rownumber,column=i)
            jsonrequest[keylist[i-1]]=cell.value
            return jsonrequest





